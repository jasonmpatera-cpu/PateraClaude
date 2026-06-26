#!/usr/bin/env python3
"""
Receipt Processor — Local Python alternative.

Watches a folder for receipt images/PDFs, extracts text via OCR,
and logs expenses to an Excel spreadsheet.

Usage:
    python receipt_processor.py                    # Process ~/Receipts once
    python receipt_processor.py --watch            # Watch folder continuously
    python receipt_processor.py --folder /path     # Custom folder
    python receipt_processor.py --spreadsheet out.xlsx
"""

import argparse
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import pytesseract
    from PIL import Image
except ImportError:
    print("Install dependencies: pip install -r requirements.txt")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    import fitz  # PyMuPDF
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

HEADERS = ["Date", "Vendor", "Amount", "Category", "Payment Method",
           "File Name", "File Path", "Notes"]

CATEGORY_PATTERNS = [
    (r"grocery|kroger|walmart|costco|safeway|trader joe|whole foods|aldi", "Groceries"),
    (r"restaurant|cafe|coffee|starbucks|mcdonald|burger|pizza|diner", "Dining"),
    (r"gas|shell|chevron|exxon|mobil|fuel|\bbp\b", "Gas"),
    (r"pharmacy|cvs|walgreens|rx|prescription|medical|doctor|clinic", "Healthcare"),
    (r"amazon|target|best buy|home depot|lowes|ikea", "Shopping"),
    (r"electric|water|internet|phone|utility|at&t|verizon|comcast", "Utilities"),
    (r"hotel|airbnb|airline|flight|uber|lyft", "Travel"),
]

PAYMENT_PATTERNS = [
    (r"visa", "Visa"),
    (r"mastercard|mc", "Mastercard"),
    (r"amex|american express", "Amex"),
    (r"discover", "Discover"),
    (r"debit", "Debit"),
    (r"cash", "Cash"),
    (r"apple pay", "Apple Pay"),
]


def extract_text_from_image(path: Path) -> str:
    img = Image.open(path)
    return pytesseract.image_to_string(img)


def extract_text_from_pdf(path: Path) -> str:
    if not HAS_PDF:
        print(f"  Skipping PDF {path.name} (install PyMuPDF: pip install pymupdf)")
        return ""
    doc = fitz.open(path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in (".jpg", ".jpeg", ".png", ".heic", ".tiff", ".bmp", ".webp"):
        return extract_text_from_image(path)
    elif suffix == ".pdf":
        return extract_text_from_pdf(path)
    return ""


def parse_receipt(text: str) -> dict:
    result = {"vendor": "", "amount": "", "category": "", "payment_method": ""}

    lines = [l.strip() for l in text.split("\n") if l.strip()]
    if lines:
        result["vendor"] = lines[0][:50]

    total_patterns = [
        r"(?:total|grand total|amount due|balance due|total due)\s*[:\s]*\$?\s*(\d+[.,]\d{2})",
        r"\$\s*(\d+[.,]\d{2})\s*$",
        r"(\d+[.,]\d{2})\s*(?:total|due)",
    ]
    for pattern in total_patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            result["amount"] = match.group(1).replace(",", ".")
            break

    for pattern, method in PAYMENT_PATTERNS:
        if re.search(pattern, text, re.IGNORECASE):
            result["payment_method"] = method
            break

    for pattern, category in CATEGORY_PATTERNS:
        if re.search(pattern, text, re.IGNORECASE):
            result["category"] = category
            break

    return result


def get_month_folder(base: Path) -> Path:
    now = datetime.now()
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]
    year_folder = base / str(now.year)
    month_folder = year_folder / f"{now.month:02d}-{month_names[now.month - 1]}"
    month_folder.mkdir(parents=True, exist_ok=True)
    return month_folder


def rename_receipt(path: Path, parsed: dict) -> str:
    date_str = datetime.now().strftime("%Y-%m-%d")
    vendor = re.sub(r'[^\w\s-]', '', parsed["vendor"]).strip().replace(" ", "_")[:30]
    amount = parsed["amount"] if parsed["amount"] else "unknown"
    suffix = path.suffix
    return f"{date_str}_{vendor}_${amount}{suffix}"


def init_spreadsheet(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 50
    wb.save(path)
    return wb


def log_to_spreadsheet(spreadsheet_path: Path, data: dict, file_name: str, file_path: str):
    wb = init_spreadsheet(spreadsheet_path)
    ws = wb["Expenses"]
    row = [
        datetime.now().strftime("%Y-%m-%d"),
        data["vendor"],
        float(data["amount"]) if data["amount"] else "",
        data["category"],
        data["payment_method"],
        file_name,
        file_path,
        "",
    ]
    ws.append(row)
    wb.save(spreadsheet_path)


def process_file(file_path: Path, base_folder: Path, spreadsheet_path: Path) -> bool:
    print(f"  Processing: {file_path.name}")
    text = extract_text(file_path)
    if not text.strip():
        print(f"    No text extracted, skipping")
        return False

    parsed = parse_receipt(text)
    new_name = rename_receipt(file_path, parsed)
    month_folder = get_month_folder(base_folder)
    dest = month_folder / new_name

    # Handle duplicates
    counter = 1
    while dest.exists():
        stem = dest.stem + f"_{counter}"
        dest = month_folder / (stem + dest.suffix)
        counter += 1

    shutil.move(str(file_path), str(dest))
    log_to_spreadsheet(spreadsheet_path, parsed, new_name, str(dest))

    print(f"    Vendor: {parsed['vendor']}")
    print(f"    Amount: ${parsed['amount']}")
    print(f"    Category: {parsed['category']}")
    print(f"    Saved to: {dest}")
    return True


def process_folder(folder: Path, spreadsheet_path: Path):
    extensions = {".jpg", ".jpeg", ".png", ".heic", ".tiff", ".bmp", ".webp", ".pdf"}
    files = [f for f in folder.iterdir()
             if f.is_file() and f.suffix.lower() in extensions]

    if not files:
        print("No receipt files found in inbox.")
        return

    print(f"Found {len(files)} receipt(s) to process.\n")
    processed = 0
    for f in sorted(files):
        if process_file(f, folder, spreadsheet_path):
            processed += 1

    print(f"\nDone. Processed {processed}/{len(files)} receipts.")
    print(f"Spreadsheet: {spreadsheet_path}")


def watch_folder(folder: Path, spreadsheet_path: Path, interval: int = 30):
    print(f"Watching {folder} for new receipts (every {interval}s)...")
    print("Press Ctrl+C to stop.\n")
    seen = set()
    extensions = {".jpg", ".jpeg", ".png", ".heic", ".tiff", ".bmp", ".webp", ".pdf"}

    while True:
        files = [f for f in folder.iterdir()
                 if f.is_file() and f.suffix.lower() in extensions and f.name not in seen]
        for f in files:
            seen.add(f.name)
            process_file(f, folder, spreadsheet_path)
        time.sleep(interval)


def main():
    parser = argparse.ArgumentParser(description="Process receipt images into organized folders + Excel spreadsheet")
    parser.add_argument("--folder", type=str, default=str(Path.home() / "Receipts"),
                        help="Receipt folder path (default: ~/Receipts)")
    parser.add_argument("--spreadsheet", type=str, default=None,
                        help="Excel file path (default: <folder>/expense_log.xlsx)")
    parser.add_argument("--watch", action="store_true",
                        help="Continuously watch folder for new receipts")
    parser.add_argument("--interval", type=int, default=30,
                        help="Watch interval in seconds (default: 30)")
    args = parser.parse_args()

    folder = Path(args.folder)
    folder.mkdir(parents=True, exist_ok=True)

    spreadsheet = Path(args.spreadsheet) if args.spreadsheet else folder / "expense_log.xlsx"

    print(f"Receipt folder: {folder}")
    print(f"Spreadsheet: {spreadsheet}\n")

    if args.watch:
        watch_folder(folder, spreadsheet, args.interval)
    else:
        process_folder(folder, spreadsheet)


if __name__ == "__main__":
    main()
